import openpyxl as op
import os

workbook = op.load_workbook('modelo_rco.xlsx')

# Selecionar a planilha desejada
sheet = workbook['Formulário']

# Acessar a célula principal do intervalo mesclado (por exemplo, 8A)
vet = [sheet['A8'], sheet['B15'], sheet['C15'], sheet['E15']]

# Alterar o valor da célula
vet[0].value = input("Entre com o valor da celula A8: ")

# Salvar as alterações no arquivo
workbook.save('modelo_rco.xlsx')

#Coletando o nome do usuario para preenchimento
username = os.getenv('USERNAME')


class Excel:
    finalidade
    numRequest
    numProject
    name 
    date
    user
    dateCreation

    

#workbook.save('nomes.xlsx') --- Salvar planilha
#workbook.create_sheet('Teste', 0) --- Criar uma aba nova
#print(workbook.sheetnames) --- Mostrar as abas disponiveis